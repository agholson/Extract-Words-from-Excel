"""
This program exports bold, italic, or underlined words from an Excel sheet to their own respective files.
"""
from typing import List
from openpyxl import load_workbook


def write_list_to_file(file_name: str, list_name: List[str]):
    """
    Function writes all the contents from a list to a file, overwriting the prior content.
    :param file_name: name of the file, e.g. bold_words.txt
    :param list_name: name of the list
    :return:
    """
    # Write to a file, overwriting the old contents
    file = open(file_name, 'w')

    # Loop through the list, append a newline character to each line
    for item in list_name:
        file.writelines(item + '\n')

    # Close the file
    file.close()


if __name__ == '__main__':
    # We load the contents of the work sheet here; we assume, it
    wb = load_workbook(filename='ExtractBoldText.xlsx')

    # Assume the first sheet is the one of importance. If it is not, then change it below here.
    ws = wb.worksheets[0]

    # Get the letters corresponding with the number of columns
    number_of_columns = ws.max_column  # Get the number of columns

    # Get the letter corresponding to the number, e.g. 3 = C
    last_column_letter = chr(ord('@') + number_of_columns)

    # Use this list to track the list of words we care about
    bold_word_list = []
    italic_word_list = []
    underline_word_list = []

    # Iterate through all the columns in the sheet
    for row in ws["A":last_column_letter]:
        # Iterate through all of the cells in the row
        for cell in row:
            try:
                # If the cell is bold, then save this text and print it out
                if cell.font.bold:
                    # Print the value from the cell
                    bold_word = cell.value
                    print(bold_word)

                    # Add the value to our list
                    bold_word_list.append(bold_word)

                elif cell.font.italic:
                    italic_word = cell.value

                    # Add the value to our list
                    italic_word_list.append(italic_word)

                elif cell.font.underline:
                    underline_word_list.append(cell.value)

            except AttributeError:
                pass

    # Write our special words to a file, if the list is not empty.
    if bold_word_list:
        write_list_to_file(file_name='bold_words.txt', list_name=bold_word_list)

    if italic_word_list:
        write_list_to_file(file_name='italic_words.txt', list_name=italic_word_list)

    if underline_word_list:
        write_list_to_file(file_name='underline_words.txt', list_name=underline_word_list)

    print(f'Successfully wrote {len(bold_word_list)} bold words to "bold_words.txt", {len(italic_word_list)} italic \n'
          f'words to "italic_words.txt", and {len(underline_word_list)} underlined words to "underline_words.txt".')
